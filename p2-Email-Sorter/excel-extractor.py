import re
from openpyxl import load_workbook

# Input and output file paths
input_file = 'data.xlsx'
output_file = 'emails.txt'

# Regex pattern for matching emails
email_pattern = re.compile(r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b')

emails = set()

# Load the workbook
wb = load_workbook(filename=input_file, read_only=True, data_only=True)

# Loop through all sheets and cells
for sheet in wb.worksheets:
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                found_emails = email_pattern.findall(cell)
                emails.update(email.lower() for email in found_emails)

# Save unique emails to text file
with open(output_file, 'w', encoding='utf-8') as f:
    for email in sorted(emails):
        f.write(email + '\n')

print(f"✅ Extracted {len(emails)} emails to {output_file}")
